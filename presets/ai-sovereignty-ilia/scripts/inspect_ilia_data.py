#!/usr/bin/env python3
"""Inspect ILIA Excel files for curated-stage operationalization."""

from __future__ import annotations

import argparse
from pathlib import Path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", default="data")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("openpyxl is required to inspect ILIA Excel files.")
        return 1

    for path in sorted(Path(args.data_dir).rglob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        print(f"# {path}")
        for sheet in wb.worksheets:
            rows = sheet.max_row or 0
            cols = sheet.max_column or 0
            headers = []
            if rows and cols:
                for row in sheet.iter_rows(min_row=1, max_row=min(3, rows), values_only=True):
                    headers.append([str(cell) if cell is not None else "" for cell in row[: min(cols, 8)]])
            print(f"- {sheet.title}: {rows} rows x {cols} cols")
            for header in headers:
                print("  - " + " | ".join(header))
        print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
